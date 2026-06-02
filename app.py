import streamlit as st
import anthropic
import base64
import json
import io
import re
from datetime import datetime
import plotly.graph_objects as go
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Supplier Price Comparison",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
.hdr{background:#203764;color:#fff;padding:1.25rem 1.75rem;border-radius:12px;margin-bottom:1.5rem}
.hdr h1{margin:0;font-size:1.7rem}
.hdr p{margin:.3rem 0 0;opacity:.8;font-size:.95rem}
div[data-testid="metric-container"]{background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;padding:.75rem 1rem}
</style>
""", unsafe_allow_html=True)

# ─── Session state ─────────────────────────────────────────────────────────────
for k in ('extracted', 'grouped'):
    if k not in st.session_state:
        st.session_state[k] = [] if k == 'extracted' else {}

# ─── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚙️ ตั้งค่า")
    api_key = st.text_input("🔑 Anthropic API Key", type="password",
                             help="ขอได้ที่ console.anthropic.com/keys")
    project_name = st.text_input("📌 ชื่อโปรเจกต์", placeholder="เช่น KPSxMonchhichi 2026")
    st.markdown("---")
    st.markdown("""**📋 วิธีใช้งาน**
1. ใส่ Anthropic API Key
2. ระบุชื่อโปรเจกต์
3. อัพโหลดใบเสนอราคา (PDF/รูป)
4. กด **ประมวลผล**
5. ดูผล → ดาวน์โหลด Excel → ส่ง Email

*รองรับ PDF ภาษาไทย/อังกฤษ และรูปถ่าย*""")
    st.markdown("---")
    if st.button("🗑️ ล้างข้อมูลทั้งหมด", use_container_width=True):
        st.session_state.extracted = []
        st.session_state.grouped = {}
        st.rerun()

# ─── Header ────────────────────────────────────────────────────────────────────
st.markdown("""<div class="hdr">
  <h1>📊 Supplier Price Comparison</h1>
  <p>อัพโหลดใบเสนอราคา → AI วิเคราะห์อัตโนมัติ → เปรียบเทียบราคา → ดาวน์โหลด Excel</p>
</div>""", unsafe_allow_html=True)

# ─── Helpers ────────────────────────────────────────────────────────────────────
COLORS = ['#3B82F6','#10B981','#F59E0B','#EF4444','#8B5CF6','#EC4899','#14B8A6','#F97316']

def encode_b64(f):
    f.seek(0)
    return base64.standard_b64encode(f.read()).decode('utf-8')

def mime(name):
    ext = name.lower().rsplit('.',1)[-1]
    return {'pdf':'application/pdf','png':'image/png','jpg':'image/jpeg','jpeg':'image/jpeg'}.get(ext,'image/jpeg')

EXTRACTION_PROMPT = """อ่านใบเสนอราคานี้อย่างละเอียด แล้วส่งคืน JSON เท่านั้น (ไม่มีข้อความอื่น):

{
  "product_category": "ชื่อหมวดสินค้า เช่น Paper Sleeve, กล่องใส่พวงกุญแจ, Mystery Box",
  "supplier": "ชื่อบริษัทผู้เสนอราคา",
  "ref": "เลขที่ใบเสนอราคา",
  "date": "วันที่",
  "brand": "ชื่อแบรนด์สินค้า",
  "material": "วัตถุดิบ/วัสดุ",
  "spec": "ขนาด/Specification",
  "payment": "เงื่อนไขชำระเงิน",
  "shipment": "ระยะเวลาส่งมอบ",
  "tooling": null,
  "remark": "หมายเหตุ (ถ้ามี)",
  "options": [
    {
      "option_name": "Option 1 — วานิช (หรือชื่อ supplier ถ้ามีแค่ตัวเดียว)",
      "other_detail": "รายละเอียด เช่น พิมพ์ 4 สี + อาบวานิช",
      "moq_prices": [
        {"qty": 3000, "price": 7.80},
        {"qty": 5000, "price": 5.40}
      ]
    }
  ]
}

กฎ: ถ้ามีหลาย Option (เช่น Option1/Option2 หรือมี Spot UV / ไม่มี) ให้แยกใน options[]
ราคาต้องเป็น "ราคาต่อชิ้น" เท่านั้น (ไม่ใช่ราคารวม) qty เป็นจำนวนเต็ม
ถ้าอ่านไม่ชัดใส่ null"""

def extract_one(uploaded_file, client):
    b64 = encode_b64(uploaded_file)
    mt  = mime(uploaded_file.name)
    if mt == 'application/pdf':
        content = [{"type":"document","source":{"type":"base64","media_type":"application/pdf","data":b64}},
                   {"type":"text","text":EXTRACTION_PROMPT}]
    else:
        content = [{"type":"image","source":{"type":"base64","media_type":mt,"data":b64}},
                   {"type":"text","text":EXTRACTION_PROMPT}]
    msg = client.messages.create(model="claude-sonnet-4-6", max_tokens=2048,
                                  messages=[{"role":"user","content":content}])
    raw = msg.content[0].text.strip()
    raw = re.sub(r'^```(?:json)?\n?','',raw)
    raw = re.sub(r'\n?```$','',raw)
    return json.loads(raw)

def group_by_cat(items):
    g = {}
    for item in items:
        cat = item.get('product_category','ไม่ระบุ')
        g.setdefault(cat,[]).append(item)
    return g

def build_sheet(cat, items):
    all_moqs = sorted({e['qty']
                       for it in items
                       for opt in it.get('options',[])
                       for e in opt.get('moq_prices',[])})
    sups = []
    for it in items:
        for opt in it.get('options',[]):
            mp = {str(e['qty']): e['price'] for e in opt.get('moq_prices',[]) if e.get('price') is not None}
            sups.append({
                'col_header': f"{it.get('supplier','')}\n{opt.get('option_name','')}".strip('\n'),
                'date': it.get('date',''), 'attach': it.get('ref',''),
                'product_name': it.get('product_category',''), 'ref': it.get('ref',''),
                'brand': it.get('brand',''), 'material': it.get('material',''),
                'other_detail': opt.get('other_detail',''), 'supplier': it.get('supplier',''),
                'spec': it.get('spec',''), 'payment': it.get('payment',''),
                'moq_prices': mp, 'tooling': it.get('tooling'),
                'min_order': f"{min(all_moqs):,} pcs." if all_moqs else '',
                'shipment': it.get('shipment',''), 'remark': it.get('remark',''),
            })
    return {'name': cat, 'moq_list': all_moqs, 'suppliers': sups}

def make_chart(sd):
    moqs = [int(m) for m in sd['moq_list']]
    fig  = go.Figure()
    for i, s in enumerate(sd['suppliers']):
        prices = [float(s['moq_prices'].get(m) or s['moq_prices'].get(str(m)) or 0) or None for m in moqs]
        label  = (s.get('col_header') or s.get('supplier','')).replace('\n',' — ')
        fig.add_trace(go.Bar(name=label, x=[f"{m:,} pcs" for m in moqs], y=prices,
                             marker_color=COLORS[i % len(COLORS)],
                             text=[f"฿{p:.2f}" if p else "" for p in prices],
                             textposition='outside'))
    fig.update_layout(barmode='group', plot_bgcolor='white', paper_bgcolor='white',
                      legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
                      yaxis=dict(title="ราคา (THB/pcs)", tickprefix="฿"),
                      xaxis=dict(title="MOQ"), height=380,
                      margin=dict(t=40, b=40, l=60, r=20))
    return fig

def make_excel(grouped, title):
    thin = Side(style='thin')
    BDR  = Border(top=thin, bottom=thin, left=thin, right=thin)
    DARK = PatternFill("solid", fgColor="203764")
    GRN  = PatternFill("solid", fgColor="E2EFDA")
    LBL  = PatternFill("solid", fgColor="D9D9D9")
    SUP  = PatternFill("solid", fgColor="BDD7EE")
    CTR  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LFT  = Alignment(horizontal='left', vertical='center', wrap_text=True)
    TOP  = ["Date Order","Attach Document","Product Name","Ref. Supplier","Pic.",
            "Brand","Material","Other Detail","Supplier","Spec","Payment Condition","Unit (EA)"]
    BOT  = ["Package&Acc. Cost (THB)","Total Costex.vat","Tooling Cost (THB)",
            "Min orderpcs.","Retail Price","% Margin","Production Order","Shipment","Remark"]
    KMAP = {"Date Order":"date","Attach Document":"attach","Product Name":"product_name",
            "Ref. Supplier":"ref","Brand":"brand","Material":"material","Other Detail":"other_detail",
            "Supplier":"supplier","Spec":"spec","Payment Condition":"payment",
            "Tooling Cost (THB)":"tooling","Min orderpcs.":"min_order","Shipment":"shipment","Remark":"remark"}
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    for cat, items in grouped.items():
        sd   = build_sheet(cat, items)
        ws   = wb.create_sheet(title=cat[:31])
        moqs = [int(m) for m in sd['moq_list']]
        sups = sd['suppliers']
        tc   = 1 + len(sups)
        mlbls = [f"Product Cost @ {m:,} pcs (THB)" for m in moqs]
        labels = TOP + mlbls + BOT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=tc)
        h = ws.cell(row=1, column=1)
        h.value="Comparison"; h.fill=DARK; h.font=Font(color="FFFFFF",bold=True,size=14)
        h.alignment=CTR; h.border=BDR
        for c in range(2,tc+1): ws.cell(1,c).fill=DARK; ws.cell(1,c).border=BDR
        ws.row_dimensions[1].height=28
        ws.cell(2,1).value="รายการ / Supplier →"; ws.cell(2,1).fill=LBL
        ws.cell(2,1).font=Font(bold=True,size=10); ws.cell(2,1).alignment=CTR; ws.cell(2,1).border=BDR
        for si,s in enumerate(sups,2):
            c=ws.cell(2,si); c.value=s.get('col_header',''); c.fill=SUP
            c.font=Font(bold=True,size=10); c.alignment=CTR; c.border=BDR
        ws.row_dimensions[2].height=40
        moq_rows={}
        for ri,lbl in enumerate(labels):
            rn=ri+3; is_moq=lbl.startswith("Product Cost @")
            if is_moq:
                mv=int(lbl.split("@")[1].split("pcs")[0].strip().replace(",",""))
                moq_rows[mv]=rn
            lc=ws.cell(rn,1); lc.value=lbl; lc.fill=LBL
            lc.font=Font(bold=True,size=10); lc.alignment=LFT; lc.border=BDR
            ws.row_dimensions[rn].height=80 if lbl=="Pic." else 18
            for si,s in enumerate(sups,2):
                dc=ws.cell(rn,si); dc.border=BDR; dc.alignment=CTR; dc.font=Font(size=10)
                if lbl=="Unit (EA)": dc.value="EA"
                elif is_moq:
                    p=s['moq_prices'].get(mv) or s['moq_prices'].get(str(mv))
                    if p: dc.value=float(p); dc.number_format="#,##0.00"
                else:
                    k=KMAP.get(lbl)
                    if k and s.get(k): dc.value=s[k]
        for mv,rn in moq_rows.items():
            ps=[(si,float(s['moq_prices'].get(mv) or s['moq_prices'].get(str(mv)) or 0))
                for si,s in enumerate(sups,2)]
            ps=[x for x in ps if x[1]>0]
            if ps:
                mn=min(p for _,p in ps)
                for ci,p in ps:
                    if p==mn: c=ws.cell(rn,ci); c.fill=GRN; c.font=Font(bold=True,size=10); c.number_format="#,##0.00"
        ws.column_dimensions['A'].width=32
        for c in range(2,tc+1): ws.column_dimensions[get_column_letter(c)].width=26
    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()

def make_email(grouped, project_name):
    today = datetime.now().strftime("%d/%m/%Y")
    lines = []
    for cat, items in grouped.items():
        sd = build_sheet(cat, items)
        if not sd['moq_list']: continue
        lm = sd['moq_list'][-1]
        best = min(sd['suppliers'], key=lambda s: float(s['moq_prices'].get(lm) or s['moq_prices'].get(str(lm)) or 1e9))
        bp = float(best['moq_prices'].get(lm) or best['moq_prices'].get(str(lm)) or 0)
        lines.append(f"  • {cat}: {best.get('col_header','').replace(chr(10),' ')} @ ฿{bp:.2f}/pcs (MOQ {int(lm):,} pcs)")
    warns = [f"  • {cat} ({it.get('supplier','')}): {it.get('remark','')}"
             for cat, items in grouped.items() for it in items if it.get('remark')]
    total_sups = sum(sum(len(it.get('options',[])) for it in items) for items in grouped.values())
    return f"""เรียน ทีม Marketing,

ทีมจัดซื้อได้เปรียบเทียบใบเสนอราคาสำหรับ {project_name or 'โปรเจกต์นี้'}
จาก {total_sups} ตัวเลือก Supplier ใน {len(grouped)} หมวดสินค้า เรียบร้อยแล้ว

📌 สรุปราคาที่ดีที่สุด (ณ {today}):
{chr(10).join(lines)}

📎 แนบไฟล์: Price_Comparison.xlsx
{(chr(10)+'⚠️ หมายเหตุ:'+chr(10)+chr(10).join(warns)) if warns else ''}

กรุณาแจ้งจำนวน Order ภายใน 1 สัปดาห์ เพื่อดำเนินการส่ง PO ต่อไป

ขอบคุณครับ/ค่ะ
ทีมจัดซื้อ"""

# ─── Upload & Process ─────────────────────────────────────────────────────────
st.subheader("📂 อัพโหลดใบเสนอราคา")
files = st.file_uploader(
    "เลือกไฟล์ (PDF หรือรูปภาพ) — อัพโหลดหลายไฟล์พร้อมกันได้",
    type=['pdf','png','jpg','jpeg'],
    accept_multiple_files=True,
    label_visibility="collapsed",
)
if files:
    st.caption(f"📎 เลือกแล้ว {len(files)} ไฟล์: " + ", ".join(f.name for f in files))

c1, c2 = st.columns([1,5])
go_btn = c1.button("🔍 ประมวลผล", type="primary",
                    disabled=not (files and api_key), use_container_width=True)

if not api_key:
    st.info("👈 ใส่ **Anthropic API Key** ในแถบซ้ายก่อนประมวลผล")

if go_btn and files and api_key:
    client = anthropic.Anthropic(api_key=api_key)
    extracted, errors = [], []
    prog = st.progress(0); status = st.empty()
    for i, f in enumerate(files):
        status.text(f"⏳ กำลังอ่าน: {f.name}  ({i+1}/{len(files)})")
        try:
            d = extract_one(f, client)
            d['_file'] = f.name
            extracted.append(d)
        except Exception as e:
            errors.append(f.name)
            st.warning(f"⚠️ อ่านไม่สำเร็จ: **{f.name}** — {str(e)[:120]}")
        prog.progress((i+1)/len(files))
    prog.empty(); status.empty()
    st.session_state.extracted = extracted
    st.session_state.grouped   = group_by_cat(extracted)
    ok = len(extracted)
    st.success(f"✅ ประมวลผลสำเร็จ {ok}/{len(files)} ไฟล์" + (f" (ล้มเหลว: {', '.join(errors)})" if errors else ""))
    st.rerun()

# ─── Results ──────────────────────────────────────────────────────────────────
if not st.session_state.grouped:
    st.stop()

grouped = st.session_state.grouped
title   = project_name or "Price Comparison"
today   = datetime.now().strftime("%Y%m%d")

st.markdown("---")

# Metrics
m1, m2, m3, m4 = st.columns(4)
total_opts = sum(sum(len(it.get('options',[])) for it in items) for items in grouped.values())
m1.metric("หมวดสินค้า", len(grouped))
m2.metric("ตัวเลือก Supplier", total_opts)
m3.metric("ไฟล์ที่วิเคราะห์", len(st.session_state.extracted))
m4.metric("วันที่", datetime.now().strftime("%d/%m/%Y"))

# Downloads
st.markdown("#### 📥 ดาวน์โหลด")
d1, d2, _ = st.columns([1,1,2])
xl = make_excel(grouped, title)
d1.download_button("📊 Excel (.xlsx)", data=xl,
                   file_name=f"Price_Comparison_{title.replace(' ','_')}_{today}.xlsx",
                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                   use_container_width=True)
sheets_list = [build_sheet(cat, items) for cat, items in grouped.items()]
jb = json.dumps({"title":title,"date":datetime.now().strftime("%d/%m/%Y"),"sheets":sheets_list},
                ensure_ascii=False, indent=2).encode('utf-8')
d2.download_button("📄 JSON (สำหรับ Dashboard)", data=jb,
                   file_name=f"Price_Comparison_{today}_data.json",
                   mime="application/json", use_container_width=True)

# Category tabs
st.markdown("---")
st.subheader("📊 ผลเปรียบเทียบรายหมวดสินค้า")
tabs = st.tabs(list(grouped.keys()))
for tab, (cat, items) in zip(tabs, grouped.items()):
    with tab:
        sd   = build_sheet(cat, items)
        sups = sd['suppliers']
        moqs = sd['moq_list']

        # Chart
        st.plotly_chart(make_chart(sd), use_container_width=True)

        # Best supplier
        if moqs:
            lm = moqs[-1]
            valid = [(s, float(s['moq_prices'].get(lm) or s['moq_prices'].get(str(lm)) or 1e9)) for s in sups]
            valid = [(s,p) for s,p in valid if p < 1e9]
            if valid:
                bs, bp = min(valid, key=lambda x: x[1])
                st.success(f"🏆 แนะนำ: **{bs.get('col_header','').replace(chr(10),' — ')}**  ราคา ฿{bp:,.2f}/pcs ที่ MOQ {int(lm):,} pcs")

        # Supplier detail cards
        st.markdown("**รายละเอียด Supplier**")
        for s in sups:
            with st.expander(f"📋  {s.get('col_header','').replace(chr(10),' — ')}"):
                c1, c2 = st.columns(2)
                with c1:
                    st.write(f"**Supplier:** {s.get('supplier','-')}")
                    st.write(f"**วัสดุ:** {s.get('material','-')}")
                    st.write(f"**Spec:** {s.get('spec','-')}")
                with c2:
                    st.write(f"**Ref:** {s.get('ref','-')}")
                    st.write(f"**ชำระเงิน:** {s.get('payment','-')}")
                    st.write(f"**ส่งของ:** {s.get('shipment','-')}")
                if s.get('remark'):
                    st.warning(f"⚠️ {s['remark']}")
                st.markdown("**ราคาตาม MOQ:**")
                cols = st.columns(len(moqs)) if moqs else []
                for col, m in zip(cols, moqs):
                    p = float(s['moq_prices'].get(m) or s['moq_prices'].get(str(m)) or 0)
                    col.metric(f"{int(m):,} pcs", f"฿{p:,.2f}" if p else "N/A")

# Email draft
st.markdown("---")
st.subheader("✉️ ร่างอีเมล์ถึงทีม MKT")
email_txt = make_email(grouped, title)
st.text_area("", email_txt, height=300, label_visibility="collapsed", key="email_box")
st.caption("💡 เลือกทั้งหมดใน box แล้ว Ctrl+C เพื่อ copy")
